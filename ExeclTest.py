import openpyxl as xl
from openpyxl.utils import column_index_from_string

execl = xl.load_workbook('example.xlsx')
sheet = execl.get_active_sheet()
#sheet = execl.get_sheet_by_name('Sheet1')
print(sheet.max_row)

for cellObjs in sheet['A1':'C5']: #row
	for cell in cellObjs:		  #column
		print(cell.row,cell.column,cell.value)

#print(list(sheet.columns))		
Flag = 0
Sum = 0
for columnC in list(sheet.columns)[7]:
		
		if Flag == 0:
			Flag = 1
		else:
			if columnC.value is not None:
				Sum += columnC.value #column_index_from_string(columnC.value)
		
print('2017年党费总数:',Sum)

'''
for i in range(0,sheet.max_row,1):
	for rowC in list(sheet.rows)[i]:
		print(rowC.value)
'''	
	
#for rowC in sheet.columns[2]:
#	print(rowC.value)